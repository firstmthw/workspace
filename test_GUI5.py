from PySide6.QtCore import *
from PySide6.QtGui import *
from PySide6.QtWidgets import *
from CaptureDevice import *
from CharucoCalibrator import *  
from SplineBackDetector import *
from SplineSideDetector import *

import numpy as np
from scipy.interpolate import interp1d
import matplotlib.pyplot as plt
from scipy import interpolate

import sys
import matplotlib
matplotlib.use('Qt5Agg')
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import time

from openpyxl import Workbook, load_workbook
from datetime import datetime
import os


class MainApp(QMainWindow):

    def __init__(self):
        QMainWindow.__init__(self)
        self.camera_back = CaptureDevice(0)
        self.camera_side = CaptureDevice(1)     
        self.screen_index = 1
        self.timer = QTimer()
        self.timer.timeout.connect(self.display_video_stream)
        self.timer.start(30)
        self.calibrating_back = False
        self.calibrating_side = False
        self.calibrator_back =  CharucoCalibrator(0.005)
        self.calibrator_side =  CharucoCalibrator(0.005)
        self.measuring_back = False
        self.measuring_side = False
        self.back_detector = SplineBackDetector()
        self.side_detector = SplineSideDetector()
        self.personal_tab = self.create_personal_tab() 
        self.measure_tab = self.create_measure_tab()
        self.result_tab = self.create_result_tab()
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("QTabBar::tab { min-width: 100px; min-height: 50px; }")
        self.tabs.addTab(self.personal_tab, "Personal Information")
        self.tabs.addTab(self.measure_tab, "Measure")
        self.tabs.addTab(self.result_tab, "Result")
        self.setCentralWidget(self.tabs)
        self.tabs.setTabVisible(0, True)
        self.tabs.setTabVisible(1, False)
        self.tabs.setTabVisible(2, False)
        self.setFixedHeight(400)
        self.setFixedWidth(1000)
        #self.ID_HN_box = "HN123456"


    def create_personal_tab(self):
        """Initialize widgets.
        """
        personal_tab = QTabWidget()
        self.setWindowTitle('ArUco GUI')
        self.grid_layout = QGridLayout()

        # Create the label and input box widgets
        self.title_label = QLabel("การประเมินสถานะผิดปกติของแนวกระดูกสันหลัง")

        self.name_label = QLabel("Name:")
        self.name_box = QLineEdit()
        self.name_box.setFixedSize(300, 30)

        self.admin_label = QLabel("Admin:")
        self.admin_box = QLineEdit()
        self.admin_box.setFixedSize(300, 30)

        self.birthday_label = QLabel("วัน/เดือน/ปีเกิด:")
        self.birthday_box = QCalendarWidget()
        self.birthday_box.setFixedSize(300, 200)

        self.ID_HN_label = QLabel("หมายเลข HN:")
        self.ID_HN_box = QLineEdit()
        self.ID_HN_box.setFixedSize(200, 30)

        self.pose_label = QLabel("ท่าทางในการทดสอบ:")
        self.pose_box = QLineEdit()
        self.pose_box.setFixedSize(200, 30)

        self.age_label = QLabel("อายุ(ปี):")
        self.age_box = QLineEdit()
        self.age_box.setFixedSize(200, 30)

        self.cam_far_label = QLabel("กล้องห่างจากตัว(ซม.):")
        self.cam_far_box = QLineEdit()
        self.cam_far_box.setFixedSize(200, 30)

        self.cam_height_label = QLabel("กล้องสูงจากพื้น(ซม.):")
        self.cam_height_box = QLineEdit()
        self.cam_height_box.setFixedSize(300, 30)

        self.height_label = QLabel("ส่วนสูง(ซม.):")
        self.height_box = QLineEdit()
        self.height_box.setFixedSize(200, 30)

        self.weight_label = QLabel("น้ำหนัก(กก.):")
        self.weight_box = QLineEdit()
        self.weight_box.setFixedSize(200, 30)

        self.Submit_button = QPushButton("Submit")
        self.Submit_button.clicked.connect(self.submit)

       # Add the label and input box widgets to the grid layout
        self.grid_layout.addWidget(self.name_label, 0, 0)
        self.grid_layout.addWidget(self.name_box, 0, 1)

        self.grid_layout.addWidget(self.admin_label, 1, 0)
        self.grid_layout.addWidget(self.admin_box, 1, 1)

        self.grid_layout.addWidget(self.birthday_label, 2, 0)
        self.grid_layout.addWidget(self.birthday_box, 2, 1)

        self.grid_layout.addWidget(self.ID_HN_label, 0, 2)
        self.grid_layout.addWidget(self.ID_HN_box, 0, 3)

        self.grid_layout.addWidget(self.pose_label, 1, 2)
        self.grid_layout.addWidget(self.pose_box, 1, 3)

        self.grid_layout.addWidget(self.age_label, 2, 2)
        self.grid_layout.addWidget(self.age_box, 2, 3)

        self.grid_layout.addWidget(self.cam_far_label, 2, 4)
        self.grid_layout.addWidget(self.cam_far_box, 2, 5)

        self.grid_layout.addWidget(self.cam_height_label, 3, 0)
        self.grid_layout.addWidget(self.cam_height_box, 3, 1)

        self.grid_layout.addWidget(self.height_label, 3, 2)
        self.grid_layout.addWidget(self.height_box, 3, 3)

        self.grid_layout.addWidget(self.weight_label, 3, 4)
        self.grid_layout.addWidget(self.weight_box, 3, 5)

        self.main_layout = QVBoxLayout()
        self.main_layout.addWidget(self.title_label)
        self.main_layout.addLayout(self.grid_layout)
        self.main_layout.addWidget(self.Submit_button)

        # Set the layout of the widget
        personal_tab.setLayout(self.main_layout)
        return personal_tab
   

    def create_measure_tab(self):
        """Initialize widgets.
        """
        measure_tab = QTabWidget()
        self.setWindowTitle('ArUco GUI')
        #self.setFixedSize(QSize(1000, 800))
    
        self.image_label_back = QLabel()
        self.image_label_back.setFixedSize(QSize(self.camera_back.preview_size[0], self.camera_back.preview_size[1]))

        self.image_label_side = QLabel()
        self.image_label_side.setFixedSize(QSize(self.camera_side.preview_size[0], self.camera_side.preview_size[1]))

        self.quit_button = QPushButton("Quit")
        self.quit_button.clicked.connect(self.close)

        self.switch_button = QPushButton("Switch")
        self.switch_button.clicked.connect(self.switch_screen)

        self.save_button = QPushButton("Save")
        self.save_button.clicked.connect(self.save_snapshot_back)
        self.save_button.clicked.connect(self.save_snapshot_side)

        self.calibrating_back_button = QPushButton("calibrate")
        self.calibrating_back_button.clicked.connect(self.calibrating_back_fcn)

        self.calibrating_side_button = QPushButton("calibrate")
        self.calibrating_side_button.clicked.connect(self.calibrating_side_fcn)

        self.measuring_back_button = QPushButton("measure")
        self.measuring_back_button.clicked.connect(self.measuring_back_fcn)
        
        self.measuring_side_button = QPushButton("measure")
        self.measuring_side_button.clicked.connect(self.measuring_side_fcn)

        self.back_layout = QVBoxLayout()
        self.back_layout.addWidget(self.image_label_back)
        self.back_layout.addWidget(self.calibrating_back_button)
        self.back_layout.addWidget(self.measuring_back_button)

        self.side_layout = QVBoxLayout()
        self.side_layout.addWidget(self.image_label_side)
        self.side_layout.addWidget(self.calibrating_side_button)
        self.side_layout.addWidget(self.measuring_side_button)

        self.mid_layout = QVBoxLayout()
        self.mid_layout.addWidget(self.quit_button)
        self.mid_layout.addWidget(self.switch_button)
        self.mid_layout.addWidget(self.save_button)

        self.main_layout = QHBoxLayout()
        self.main_layout.addLayout(self.back_layout)       
        self.main_layout.addLayout(self.side_layout)
        self.main_layout.addLayout(self.mid_layout)
        
        measure_tab.setLayout(self.main_layout)
        return measure_tab


    def create_result_tab(self):
        result_tab = QTabWidget()
        #self.setFixedSize(QSize(1000, 800))      

        self.quit_button = QPushButton("Quit")
        self.quit_button.clicked.connect(self.close)

        self.clear_graph_button = QPushButton("clear")
        self.clear_graph_button.clicked.connect(self.clear_graph)
    
        self.back_figure = plt.Figure(figsize = (3.6,6.4))
        self.back_ax = self.back_figure.add_subplot(111)
        self.back_canvas = FigureCanvas(self.back_figure)
        #change the size of the figure
        #self.back_canvas.setMinimumSize(400, 200)

        self.side_figure = plt.Figure()
        self.side_ax = self.side_figure.add_subplot(111)
        self.side_canvas = FigureCanvas(self.side_figure)
        #change the size of the figure
        #self.side_canvas.setMinimumSize(400, 200)

        self.back_x_label = QLabel()
        self.back_y_label = QLabel()
        self.back_z_label = QLabel()

        self.side_angle = QLabel()
        self.side_shoulder_angle = QLabel()
        self.side_neck_angle = QLabel()

        self.back_result_layout = QVBoxLayout()
        self.back_result_layout.addWidget(self.back_canvas)
        self.back_result_layout.addWidget(self.back_x_label)
        self.back_result_layout.addWidget(self.back_y_label)
        self.back_result_layout.addWidget(self.back_z_label)

        self.side_result_layout = QVBoxLayout()
        self.side_result_layout.addWidget(self.side_canvas)
        self.side_result_layout.addWidget(self.side_angle)
        self.side_result_layout.addWidget(self.side_shoulder_angle)
        self.side_result_layout.addWidget(self.side_neck_angle)

        self.result_layout = QHBoxLayout()
        self.result_layout.addLayout(self.back_result_layout)       
        self.result_layout.addLayout(self.side_result_layout)
        self.result_layout.addWidget(self.quit_button)
        #self.result_layout.addWidget(self.clear_graph_button)

        result_tab.setLayout(self.result_layout)
        return result_tab
        
    def clear_graph(self):
        self.back_ax.cla()
        self.back_ax.plot([])            
        self.back_canvas.draw()

    def submit(self):
        flist = os.listdir()
        if 'data_%s.xlsx'%self.ID_HN_box.text() in flist:
            self.wb = load_workbook('data_%s.xlsx'%self.ID_HN_box.text())
            print('file exist, open file')

            # show the other tabs
            self.tabs.setTabVisible(0, False)
            self.tabs.setTabVisible(1, True)
            self.tabs.setTabVisible(2, True)
            self.tabs.setCurrentWidget(self.measure_tab)
            self.setFixedHeight(800)
            self.setFixedWidth(1000)

        else:
            self.wb = load_workbook('data_template.xlsx')
            print('file not exist, create new file')

            # update overview
            self.overview_sheet = self.wb['ข้อมูลด้านหลัง']
            self.overview_sheet['B2'] = self.name_box.text()
            self.overview_sheet['F2'] = self.ID_HN_box.text()
            self.overview_sheet['F3'] = self.pose_box.text()
            self.overview_sheet['B3'] = self.admin_box.text()
            self.overview_sheet['B5'] = self.birthday_box.selectedDate().toString("yyyy/MM/dd")
            self.overview_sheet['D5'] = self.age_box.text()
            self.overview_sheet['F5'] = self.cam_far_box.text()
            self.overview_sheet['B6'] = self.height_box.text()
            self.overview_sheet['D6'] = self.weight_box.text()
            self.overview_sheet['F6'] = self.cam_height_box.text()

            topic = [self.name_box.text(),self.ID_HN_box.text(),self.admin_box.text(),self.birthday_box.selectedDate().toString("yyyy/MM/dd"),self.age_box.text(),self.height_box.text(),self.weight_box.text(),self.pose_box.text(),self.cam_far_box.text(),self.cam_height_box.text()]

            if any(key == "" for key in topic):
                print("NO IS NO")
                #self.alert_box.exec()
            else:
            # show the other tabs
                self.tabs.setTabVisible(0, False)
                self.tabs.setTabVisible(1, True)
                self.tabs.setTabVisible(2, True)
                self.tabs.setCurrentWidget(self.measure_tab)

            self.wb.save('data_%s.xlsx'%self.ID_HN_box.text())  

    def display_video_stream(self):
        """Read frame from camera and repaint QLabel widget.
        """
        self.camera_side.read()
        self.camera_back.read()

        frame_side = self.camera_side.preview()
        frame_back = self.camera_back.preview()

       
        if self.screen_index == 1:
            image_back = QImage(frame_back, frame_back.shape[1], frame_back.shape[0], frame_back.strides[0], QImage.Format_RGB888)
            self.image_label_back.setPixmap(QPixmap.fromImage(image_back))

            image_side = QImage(frame_side, frame_side.shape[1], frame_side.shape[0], frame_side.strides[0], QImage.Format_RGB888)
            self.image_label_side.setPixmap(QPixmap.fromImage(image_side))

        else:
            image_back = QImage(frame_back, frame_back.shape[1], frame_back.shape[0], frame_back.strides[0], QImage.Format_RGB888)
            self.image_label_side.setPixmap(QPixmap.fromImage(image_back))

            image_side = QImage(frame_side, frame_side.shape[1], frame_side.shape[0], frame_side.strides[0], QImage.Format_RGB888)
            self.image_label_back.setPixmap(QPixmap.fromImage(image_side))

        if self.calibrating_back == True:
            num_cal_image_back = self.calibrator_back.append(self.camera_back.raw_image())
            print(num_cal_image_back)
            if num_cal_image_back >= 20:
                self.calibrating_back = False
                self.calibrator_back.calibrate()

                
        if self.calibrating_side == True:
            num_cal_image_side = self.calibrator_side.append(self.camera_side.raw_image())
            print(num_cal_image_side)
            if num_cal_image_side >= 20:
                self.calibrating_side = False
                self.calibrator_side.calibrate()

        

        if self.measuring_back == True:
            m_back = self.calibrator_back.measure(self.camera_back.raw_image(),self.back_detector.get_ids())  
            if m_back is not None:
                self.measuring_back = False
                for i in range(len(m_back["ids"])):
                    print(m_back["ids"][i],m_back["tvecs"][i][0][0][0:3])
                #print (m_back["tvecs"][0][0][0][0:2])

                #เรียงตาม id ที่ตรวจจับได้
                self.id_new = []
                for id in self.back_detector.get_ids():
                    for i in range(len(m_back["ids"])):
                        if id == m_back["ids"][i]:
                            #หาindex ของ id จาก m_back
                            index_m_back = m_back["ids"].index(id)
                            #นำ index ไปหา tvec ของ id นั้น
                            tvec = m_back["tvecs"][index_m_back][0][0][0:3]
                            #เก็บ id และ tvec ไว้ใน list
                            self.id_new.append([id,tvec])
                print(self.id_new)

                #เรียกใช้ id และ tvec ที่เก็บไว้ใน list
                x = []
                y = []
                z = []
                #หา ค่า x,y,z ของ id ที่ตรวจจับได้
                for i in range(len(self.id_new)):
                    x.append(round(self.id_new[i][1][0],4))
                    y.append(round(self.id_new[i][1][1],4))
                    z.append(round(self.id_new[i][1][2],4))

                    
                print("x",x)
                print("y",y)
                print("z",z)

                self.back_x_label.setText("x = " + str(x))
                self.back_y_label.setText("y = " + str(y))
                self.back_z_label.setText("z = " + str(z))
                

                #ทำการเรียก spline
                tck,u = interpolate.splprep([x,y],k=3 ,s=0)
                u_new = np.linspace(u.min(),u.max(),50)
                x_new,y_new = interpolate.splev(u_new,tck,der=0)

                #print(tck[1])
                
                self.back_ax.plot(x_new,-y_new , label = 'spline')
                #หาค่าเฉลี่ยของ x,y
                x_mean = sum(x)/len(x)
                y_mean = sum(y)/len(y)
                #กำหนดขอบเขตกราฟ
                self.back_ax.set_xlim(x_mean-0.1,x_mean+0.1)
                self.back_ax.set_ylim(-y_mean-0.1,-y_mean+0.1)
                
                #show text a b c d on point coordinate
                for i in range(len(x)):
                    name_text_back = ["T1","T3","T5","T7","T9","T11"]
                    self.back_ax.plot(x[i],-y[i],'ro')
                    self.back_ax.text(x[i],-y[i],name_text_back[i],fontsize=10)
                self.back_ax.legend()

                #show graph
                self.back_canvas.draw()
                             
        
        if self.measuring_side == True:
            m_side = self.calibrator_side.measure(self.camera_side.raw_image(),self.side_detector.get_ids())
            if m_side is not None:
                self.measuring_side = False
                for i in range(len(m_side["ids"])):
                    print(m_side["ids"][i],m_side["tvecs"][i][0][0][0:2])
                #print (m_side["tvecs"][0][0][0][0:2])          
                
                #เรียงตาม id ที่ตรวจจับได้
                self.id_new = []
                for id in self.side_detector.get_ids():
                    for i in range(len(m_side["ids"])):
                        if id == m_side["ids"][i]:
                            #หาindex ของ id จาก m_back
                            index_m_side = m_side["ids"].index(id)
                            #นำ index ไปหา tvec ของ id นั้น
                            tvec = m_side["tvecs"][index_m_side][0][0][0:2]
                            #เก็บ id และ tvec ไว้ใน list
                            self.id_new.append([id,tvec])
                            
                print(self.id_new)

              

                #เรียกใช้ id และ tvec ที่เก็บไว้ใน list
                x = []
                y = []
                #หา ค่า x,y ของ id ที่ตรวจจับได้
                for i in range(len(self.id_new)):
                    x.append(self.id_new[i][1][0])
                    y.append(self.id_new[i][1][1])
                    
                print("x",x)
                print("y",y)
                

                #ให้ y เป็นค่าลบ
                y = [-i for i in y]

                self.side_ax.plot(x,y , 'o-' , label = 'spline')
                #self.side_ax.set_xlim([0.1,-0.2])

                #show text on point coordinate in graph
                for i in range(len(x)):
                    name_text_side = ["ears","neck","shoulder"]
                    self.side_ax.plot(x[i],y[i],'ro')
                    self.side_ax.text(x[i],y[i],name_text_side[i],fontsize=10)

                #show graph
                self.side_canvas.draw()
            

                degree = self.side_detector.measure_angle(self.id_new[0][1],self.id_new[1][1],self.id_new[2][1])

                print("Angle",degree)
                self.side_angle.setText('angle = '+str(degree))

                #หามุมที่คอและไหล่
                shoulder_angle, neck_angle = self.side_detector.find_angle(self.id_new[1][1],self.id_new[2][1],degree)
                self.side_shoulder_angle.setText('Shoulder Angle = '+str(shoulder_angle))
                self.side_neck_angle.setText('Neck Angle = '+str(neck_angle))

                print("Shoulder Angle",shoulder_angle)
                print("Neck Angle",neck_angle)
        
    

    def switch_screen(self):
        if self.screen_index == 1:
            self.screen_index = 2
        else:
            self.screen_index = 1

    def save_snapshot_back(self):
        """Save snapshot.
        """
        cv2.imwrite("snapshot_back.png", self.camera_back.raw_image())

    def save_snapshot_side(self):
        """Save snapshot.
        """
        cv2.imwrite("snapshot_side.png", self.camera_side.raw_image())

    def calibrating_back_fcn(self):
        self.calibrating_back = True
        print ("calibrating back")

    def calibrating_side_fcn(self):
        self.calibrating_side = True
        print ("calibrating side")

    def measuring_back_fcn(self):
        self.measuring_back = True
        print ("measuring back")

    def measuring_side_fcn(self):
        self.measuring_side = True
        print ("measuring side")

    # สร้างฟังก์ชันในการสร้างตาราง
    def create_table(self, id, lable):
        id = [str(x)for x in id]
        lable.setRowCount(1)
        lable.setColumnCount(len(id))



if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = MainApp()
    #win.resize (1000, 800) #เดิมแนวตั้ง1420, 600
    win.show()
    sys.exit(app.exec())

